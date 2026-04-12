"""
Canadian Classroom Exchange — Unit Generation Script
=====================================================
Reads content spec from Excel, generates full 7-lesson units via Claude API,
writes markdown files, and updates status back to the spreadsheet.

Usage:
    python generate_units.py --batch 1                # run all pending in Batch 1
    python generate_units.py --batch 1 --dry-run      # preview without API calls
    python generate_units.py --batch 2 --limit 3      # run first 3 pending in Batch 2
"""

import anthropic
import argparse
from dotenv import load_dotenv
load_dotenv()
import json
import os
import re
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font

# ── Config ──────────────────────────────────────────────────────────────────
EXCEL_PATH   = "canadian_classroom_content_batches.xlsx"
OUTPUT_DIR   = Path("generated_units")
MODEL        = "claude-opus-4-5"
MAX_TOKENS   = 8000

SHEET_MAP = {
    "1": "Batch 1 — Pilot",
    "2": "Batch 2 — Stress Test",
    "3": "Batch 3 — Scale",
}

# Status colours (match spreadsheet)
FILL_GENERATED = PatternFill("solid", fgColor="FFF2CC")
FILL_REGEN     = PatternFill("solid", fgColor="EAD1DC")

# Column indices (1-based, matching spreadsheet)
COL = {
    "batch":       1,
    "grade":       2,
    "subject":     3,
    "strand":      4,
    "resource_type": 5,
    "unit_theme":  6,
    "curriculum_code": 7,
    "duration":    8,
    "priority":    9,
    "notes":       10,
    "pm_comments": 11,
    "status":      12,
}

# ── Prompt template ─────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are an expert Ontario elementary curriculum designer with 15 years of classroom experience.
You create engaging, curriculum-aligned educational resources for K-3 Ontario teachers.
You deeply understand the 2020 Ontario Mathematics curriculum, the 2023 Language (Science of Reading) update,
the 2022 Science and Technology curriculum, and the 2023 Social Studies (Indigenous perspectives) update.

You always write in a warm, practical teacher voice. Your lesson plans are immediately usable —
no vague instructions, every teacher script is written out in full.

Output ONLY valid JSON. No markdown fences, no preamble, no explanation outside the JSON."""

def build_prompt(row: dict) -> str:
    return f"""Create a complete, ready-to-sell educational unit for Ontario teachers.

INPUT PARAMETERS:
- Grade: {row['grade']}
- Subject: {row['subject']}
- Strand: {row['strand']}
- Resource Type: {row['resource_type']}
- Suggested Theme: "{row['unit_theme']}" (you may refine this name)
- Curriculum Code: {row['curriculum_code']}
- Duration: {row['duration']} days
- Priority Gap This Fills: {row['notes']}

INSTRUCTIONS:
1. Infer the best structure for this strand and grade — do NOT force a rigid template.
   A coding unit might use a "Mission Briefing" instead of "Minds On". A phonics unit
   might have 4 short activities instead of 3 long ones. Match the format to the content.
2. Every lesson must have a teacher script written in full — no "[discuss with students]" placeholders.
3. Worksheets must have [IMAGE_PLACEHOLDER] tags with precise descriptions so an image
   generation script knows exactly what to create.
4. The unit must be internally coherent — later lessons build on earlier ones.
5. Use Canadian contexts throughout (Canadian coins, Canadian animals, Canadian places).
6. The descriptive_title is separate from the thematic_title — it is used for search tags only.

OUTPUT FORMAT (return this exact JSON structure):
{{
  "thematic_title": "Catchy theme name like 'Our Classroom Market'",
  "descriptive_title": "Grade 1 Mathematics: Financial Literacy — Canadian Coins",
  "grade": "{row['grade']}",
  "subject": "{row['subject']}",
  "strand": "{row['strand']}",
  "curriculum_code": "{row['curriculum_code']}",
  "resource_type": "{row['resource_type']}",
  "duration_days": {row['duration']},
  "unit_overview": "2-3 sentence description for the marketplace listing",
  "learning_goals": ["goal 1", "goal 2", "goal 3"],
  "materials_needed": ["item 1", "item 2"],
  "assessments_included": {{
    "diagnostic": "brief description",
    "formative": "brief description",
    "summative": "brief description"
  }},
  "lessons": [
    {{
      "lesson_number": 1,
      "lesson_title": "The Hook title",
      "duration_minutes": 45,
      "learning_intention": "We are learning to...",
      "success_criteria": ["I can...", "I can..."],
      "minds_on": {{
        "duration_minutes": 10,
        "hook": "Opening hook sentence for teacher to say aloud",
        "teacher_script": "Full script written out. 'Say to students: ...' format.",
        "materials": ["item"]
      }},
      "action": {{
        "duration_minutes": 25,
        "activity_name": "Name of the activity",
        "instructions": "Step-by-step instructions written for teacher.",
        "differentiation": "How to support struggling students / extend for advanced.",
        "formative_check": "What to look for / quick assessment during activity."
      }},
      "consolidation": {{
        "duration_minutes": 10,
        "exit_ticket": "Specific exit ticket question or task.",
        "discussion_prompt": "Closing discussion question."
      }},
      "worksheet": {{
        "worksheet_title": "The [Theme] Activity Sheet",
        "instructions_to_student": "Written instructions printed on the sheet",
        "parts": [
          {{
            "part_number": 1,
            "part_title": "Part 1: [Name]",
            "task_description": "What the student does",
            "image_placeholders": [
              {{
                "id": "L1_P1_IMG1",
                "description": "Precise description for image generation: style, content, size. E.g. 'B&W line art coloring-book style, Canadian nickel (5-cent coin) showing beaver on reverse, 80x80px, clean outlines'",
                "placement": "left of question text"
              }}
            ],
            "student_response_type": "circle | write | draw | colour | match | cut-and-sort",
            "answer_key": "correct answer(s)"
          }}
        ]
      }}
    }}
  ],
  "assessment_suite": {{
    "diagnostic_tracker": {{
      "title": "Day 1 Diagnostic: [Name]",
      "teacher_tip": "Tip for using this during circulating",
      "columns": ["Student Name", "Skill 1", "Skill 2", "Skill 3", "Notes"]
    }},
    "formative_trackers": [
      {{
        "lesson": 2,
        "title": "Formative Tracker — Lesson 2",
        "columns": ["Student Name", "Observation 1", "Observation 2", "Notes"]
      }}
    ],
    "summative_rubric": {{
      "title": "Summative Assessment Rubric",
      "criteria": [
        {{
          "name": "Criterion name",
          "level_1": "Description",
          "level_2": "Description",
          "level_3": "The Goal — description",
          "level_4": "Description"
        }}
      ]
    }},
    "summative_task": {{
      "title": "The [Theme] Final Challenge",
      "format": "interview | written | performance task",
      "script_or_instructions": "Full teacher script or student instructions"
    }}
  }},
  "marketplace_metadata": {{
    "descriptive_title": "Same as top-level descriptive_title — used as search tag",
    "suggested_price_cad": 9.99,
    "tags": ["financial literacy", "Grade 1", "Ontario curriculum", "Canadian coins"],
    "hidden_keywords": ["F1.1", "subitizing", "number sense"],
    "what_is_included": ["7 lesson plans", "7 worksheets", "assessment suite", "rubric"],
    "target_buyer": "Grade 1 Ontario classroom teacher"
  }}
}}"""

# ── Markdown renderer ────────────────────────────────────────────────────────
def unit_to_markdown(unit: dict, source_row: dict) -> str:
    """Convert JSON unit to structured markdown file."""
    lines = []

    # ── Header ──
    lines += [
        f"# {unit['thematic_title']}",
        f"",
        f"> **Descriptive title (search tag):** {unit['descriptive_title']}",
        f"",
        f"| Field | Value |",
        f"|---|---|",
        f"| Grade | {unit['grade']} |",
        f"| Subject | {unit['subject']} |",
        f"| Strand | {unit['strand']} |",
        f"| Curriculum code | `{unit['curriculum_code']}` |",
        f"| Resource type | {unit['resource_type']} |",
        f"| Duration | {unit['duration_days']} days |",
        f"| Generated | {datetime.now().strftime('%Y-%m-%d')} |",
        f"",
        f"## Unit overview",
        f"",
        unit['unit_overview'],
        f"",
        f"## Learning goals",
        f"",
    ]
    for goal in unit.get('learning_goals', []):
        lines.append(f"- {goal}")

    lines += [
        f"",
        f"## Materials needed",
        f"",
    ]
    for mat in unit.get('materials_needed', []):
        lines.append(f"- {mat}")

    a = unit.get('assessments_included', {})
    lines += [
        f"",
        f"## Assessments included",
        f"",
        f"- **Diagnostic:** {a.get('diagnostic','')}",
        f"- **Formative:** {a.get('formative','')}",
        f"- **Summative:** {a.get('summative','')}",
        f"",
        f"---",
        f"",
    ]

    # ── Lessons ──
    for lesson in unit.get('lessons', []):
        n = lesson['lesson_number']
        lines += [
            f"## Lesson {n}: {lesson['lesson_title']}",
            f"",
            f"**Grade:** {unit['grade']} | "
            f"**Strand:** {unit['strand']} | "
            f"**Duration:** {lesson['duration_minutes']} minutes",
            f"",
            f"**Learning intention:** {lesson['learning_intention']}",
            f"",
            f"**Success criteria:**",
            f"",
        ]
        for sc in lesson.get('success_criteria', []):
            lines.append(f"- {sc}")

        mo = lesson.get('minds_on', {})
        lines += [
            f"",
            f"### Minds On ({mo.get('duration_minutes', '')} min)",
            f"",
            f"**Hook:** {mo.get('hook', '')}",
            f"",
            f"**Teacher script:**",
            f"",
            mo.get('teacher_script', ''),
            f"",
        ]

        ac = lesson.get('action', {})
        lines += [
            f"### Action ({ac.get('duration_minutes', '')} min) — {ac.get('activity_name', '')}",
            f"",
            ac.get('instructions', ''),
            f"",
            f"**Differentiation:** {ac.get('differentiation', '')}",
            f"",
            f"**Formative check:** {ac.get('formative_check', '')}",
            f"",
        ]

        con = lesson.get('consolidation', {})
        lines += [
            f"### Consolidation ({con.get('duration_minutes', '')} min)",
            f"",
            f"**Exit ticket:** {con.get('exit_ticket', '')}",
            f"",
            f"**Discussion prompt:** {con.get('discussion_prompt', '')}",
            f"",
        ]

        # ── Worksheet ──
        ws = lesson.get('worksheet', {})
        if ws:
            lines += [
                f"---",
                f"",
                f"### Student worksheet — {ws.get('worksheet_title', '')}",
                f"",
                f"**Instructions:** {ws.get('instructions_to_student', '')}",
                f"",
            ]
            for part in ws.get('parts', []):
                lines += [
                    f"#### {part.get('part_title', '')}",
                    f"",
                    part.get('task_description', ''),
                    f"",
                    f"*Response type: {part.get('student_response_type', '')}*",
                    f"",
                    f"**Answer key:** {part.get('answer_key', '')}",
                    f"",
                ]
                for img in part.get('image_placeholders', []):
                    lines += [
                        f"```image-placeholder",
                        f"id: {img['id']}",
                        f"description: {img['description']}",
                        f"placement: {img['placement']}",
                        f"```",
                        f"",
                    ]

        lines += [f"---", f""]

    # ── Assessment suite ──
    suite = unit.get('assessment_suite', {})
    lines += [f"## Assessment suite", f""]

    diag = suite.get('diagnostic_tracker', {})
    if diag:
        lines += [
            f"### {diag.get('title', 'Diagnostic tracker')}",
            f"",
            f"*Teacher tip: {diag.get('teacher_tip', '')}*",
            f"",
        ]
        cols = diag.get('columns', [])
        if cols:
            lines.append("| " + " | ".join(cols) + " |")
            lines.append("| " + " | ".join(["---"] * len(cols)) + " |")
            lines.append("| " + " | ".join([""] * len(cols)) + " |")
        lines.append("")

    for ft in suite.get('formative_trackers', []):
        lines += [
            f"### {ft.get('title', '')}",
            f"",
        ]
        cols = ft.get('columns', [])
        if cols:
            lines.append("| " + " | ".join(cols) + " |")
            lines.append("| " + " | ".join(["---"] * len(cols)) + " |")
            lines.append("| " + " | ".join([""] * len(cols)) + " |")
        lines.append("")

    rubric = suite.get('summative_rubric', {})
    if rubric:
        lines += [f"### {rubric.get('title', 'Summative rubric')}", f""]
        lines.append("| Criteria | Level 1 | Level 2 | Level 3 (Goal) | Level 4 |")
        lines.append("|---|---|---|---|---|")
        for c in rubric.get('criteria', []):
            lines.append(
                f"| {c.get('name','')} | {c.get('level_1','')} | "
                f"{c.get('level_2','')} | {c.get('level_3','')} | {c.get('level_4','')} |"
            )
        lines.append("")

    task = suite.get('summative_task', {})
    if task:
        lines += [
            f"### {task.get('title', 'Summative task')}",
            f"",
            f"**Format:** {task.get('format', '')}",
            f"",
            task.get('script_or_instructions', ''),
            f"",
        ]

    # ── Marketplace metadata ──
    meta = unit.get('marketplace_metadata', {})
    lines += [
        f"---",
        f"",
        f"## Marketplace metadata",
        f"",
        f"```yaml",
        f"thematic_title: \"{unit['thematic_title']}\"",
        f"descriptive_title: \"{meta.get('descriptive_title', unit['descriptive_title'])}\"",
        f"suggested_price_cad: {meta.get('suggested_price_cad', 9.99)}",
        f"tags: {json.dumps(meta.get('tags', []))}",
        f"hidden_keywords: {json.dumps(meta.get('hidden_keywords', []))}",
        f"what_is_included: {json.dumps(meta.get('what_is_included', []))}",
        f"target_buyer: \"{meta.get('target_buyer', '')}\"",
        f"```",
        f"",
    ]

    return "\n".join(lines)


# ── Excel helpers ────────────────────────────────────────────────────────────
def load_pending_rows(excel_path: str, sheet_name: str) -> list[dict]:
    wb  = openpyxl.load_workbook(excel_path)
    ws  = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=False):
        status_cell = row[COL["status"] - 1]
        status = str(status_cell.value or "").lower().strip()
        if status == "pending":
            rows.append({
                "row_number":      row[0].row,
                "batch":           row[COL["batch"] - 1].value,
                "grade":           row[COL["grade"] - 1].value,
                "subject":         row[COL["subject"] - 1].value,
                "strand":          row[COL["strand"] - 1].value,
                "resource_type":   row[COL["resource_type"] - 1].value,
                "unit_theme":      row[COL["unit_theme"] - 1].value,
                "curriculum_code": row[COL["curriculum_code"] - 1].value,
                "duration":        row[COL["duration"] - 1].value or 7,
                "priority":        row[COL["priority"] - 1].value,
                "notes":           row[COL["notes"] - 1].value or "",
            })
    return rows


def update_status(excel_path: str, sheet_name: str,
                  row_number: int, status: str, comment: str = ""):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]

    status_cell  = ws.cell(row=row_number, column=COL["status"])
    comment_cell = ws.cell(row=row_number, column=COL["pm_comments"])

    status_cell.value = status
    status_cell.font  = Font(name="Arial", bold=True, size=9)

    fill_map = {
        "generated":  FILL_GENERATED,
        "needs_regen": PatternFill("solid", fgColor="EAD1DC"),
    }
    if status in fill_map:
        status_cell.fill = fill_map[status]

    if comment:
        existing = comment_cell.value or ""
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        comment_cell.value = f"{existing}\n[{ts}] {comment}".strip()

    wb.save(excel_path)


# ── Core generation ──────────────────────────────────────────────────────────
def generate_unit(client: anthropic.Anthropic, row: dict,
                  dry_run: bool = False) -> tuple[dict | None, str]:
    """
    Returns (unit_dict, error_message).
    unit_dict is None on failure.
    """
    if dry_run:
        print(f"  [DRY RUN] Would generate: {row['unit_theme']} ({row['grade']} · {row['strand']})")
        return None, "dry_run"

    print(f"  Generating: {row['unit_theme']} ({row['grade']} · {row['strand']})...")

    try:
        response = client.messages.create(
            model=MODEL,
            max_tokens=MAX_TOKENS,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": build_prompt(row)}],
        )
        raw = response.content[0].text.strip()

        # Strip accidental markdown fences
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)

        unit = json.loads(raw)
        return unit, ""

    except json.JSONDecodeError as e:
        return None, f"JSON parse error: {e}"
    except anthropic.APIError as e:
        return None, f"API error: {e}"
    except Exception as e:
        return None, f"Unexpected error: {e}"


def slug(text: str) -> str:
    """Convert text to a safe filename slug."""
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")[:60]


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch",   required=True, choices=["1", "2", "3"])
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--limit",   type=int, default=None,
                        help="Max units to generate this run (default: all pending)")
    args = parser.parse_args()

    sheet_name = SHEET_MAP[args.batch]
    OUTPUT_DIR.mkdir(exist_ok=True)
    batch_dir = OUTPUT_DIR / f"batch_{args.batch}"
    batch_dir.mkdir(exist_ok=True)

    client   = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
    rows     = load_pending_rows(EXCEL_PATH, sheet_name)

    if not rows:
        print(f"No pending rows found in '{sheet_name}'. Nothing to do.")
        return

    if args.limit:
        rows = rows[:args.limit]

    print(f"\n{'='*60}")
    print(f"Batch {args.batch} — {sheet_name}")
    print(f"Units to generate: {len(rows)}")
    print(f"Dry run: {args.dry_run}")
    print(f"{'='*60}\n")

    success = 0
    failed  = 0

    for i, row in enumerate(rows, 1):
        print(f"[{i}/{len(rows)}] {row['grade']} · {row['subject']} · {row['strand']}")

        unit, error = generate_unit(client, row, dry_run=args.dry_run)

        if args.dry_run:
            continue

        if unit is None:
            print(f"  ✗ Failed: {error}")
            update_status(EXCEL_PATH, sheet_name, row["row_number"],
                          "needs_regen", f"Generation failed: {error}")
            failed += 1
            continue

        # Save JSON (raw, for debugging + future pipeline use)
        json_path = batch_dir / f"{slug(row['unit_theme'])}.json"
        json_path.write_text(json.dumps(unit, indent=2, ensure_ascii=False))

        # Save Markdown (for PM review in Google Docs)
        md_path = batch_dir / f"{slug(row['unit_theme'])}.md"
        md_path.write_text(unit_to_markdown(unit, row))

        # Update spreadsheet
        update_status(
            EXCEL_PATH, sheet_name, row["row_number"],
            "generated",
            f"Files: {md_path.name} | Price: ${unit.get('marketplace_metadata', {}).get('suggested_price_cad', '?')} CAD"
        )

        print(f"  ✓ Saved: {md_path.name}")
        success += 1

        # Polite delay between API calls
        if i < len(rows):
            time.sleep(1.5)

    print(f"\n{'='*60}")
    print(f"Done. Success: {success} | Failed: {failed}")
    print(f"Output folder: {batch_dir.resolve()}")
    print(f"{'='*60}\n")
    if success:
        print("Next step: Share the markdown files with your PM for review.")
        print("PM updates Status column in the spreadsheet: approved / needs_regen")


if __name__ == "__main__":
    main()
